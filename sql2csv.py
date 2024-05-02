#!/usr/bin/env python3

# TODO: fix ruff
# ruff: noqa

"""
SQL to CSV.

source at
https://github.com/entorb/sql2csv

## Features
* connect to a database
* read all .sql files of current directory
* execute one after the other
* export results set as text (.csv) and Excel (.xlsx)

## Supported Databases
* PostgreSQL
* Oracle
* MS SQL
* SQLite3

## TODOs
- [x] scan SQL for dangerous commands like DROP/DELETE (incomplete!)
- [x] Limits the max number of returned rows via limit on cells = columns * rows
- [x] hashing of SQL files to prevent modification
- [ ] Excel: autosize column width
- [ ] [sqlparse](https://sqlparse.readthedocs.io/en/latest/api/) to remove SQL comments

## **SECURITY WARNING:** Only use read-only db-user accounts!
example for PostgreSQL<br/>
`GRANT SELECT ON ALL TABLES IN SCHEMA schema_name TO username`<br/>
`GRANT USAGE ON SCHEMA schema_name TO username`

## Requirements
### Oracle
Oracle Instant Client - Basic Light Package
from
https://www.oracle.com/database/technologies/instant-client/winx64-64-downloads.html
download, unzip and add dir to path

### MS SQL
Microsoft ODBC Driver for SQL Server
from
https://docs.microsoft.com/en-us/sql/connect/odbc/download-odbc-driver-for-sql-server?view=sql-server-ver15
install

"""

# Convert to .exe via
# pyinstaller --onefile --console sql2csv.py
import csv  # for .csv writing
import datetime
import decimal
import hashlib  # for sha256 checksums
import re
import sqlite3
from pathlib import Path

import cx_Oracle  # driver for Oracle - pip install cx_Oracle
import openpyxl  # for Excel writing
import psycopg2  # PostgreSQL - pip install psycopg2
import pyodbc  # ODBC driver for MS SQL and others  - pip install pyodbc

from sql2csv_credentials import credentials, hash_salt

# DB drivers
# read my credential file

# IDEA: move settings to .ini
cnt_max_cells = 100000
csv_format_datetime = "%Y-%m-%d_%H:%M:%S"
csv_format_date = "%Y-%m-%d"
csv_delimiter = "\t"  # \t ; ,
csv_quotechar = '"'
csv_newline = "\n"


#
# 1. database functions
#


def connect(*, verbose: bool = True) -> tuple:
    """
    Connect to the database server.
    """
    connection = None
    cursor = None

    if credentials["db_type"] == "postgres":
        connection = psycopg2.connect(
            host=credentials["host"],
            port=credentials["port"],
            database=credentials["database"],
            user=credentials["user"],
            password=credentials["password"],
        )

    elif credentials["db_type"] == "sqlite3":
        # here database is the path to the database file
        connection = sqlite3.connect(credentials["database"])
        credentials["host"] = "localhost"

    elif credentials["db_type"] == "oracle":
        # connection = pyodbc.connect(
        #     f"DRIVER={{ORACLE ODBC DRIVER}};SERVER=tcp:{credentials['host']},{credentials['port']};DATABASE={credentials['database']};UID={credentials['user']};PWD={credentials['password']}"  # noqa: E501
        # )
        connection = cx_Oracle.connect(
            credentials["user"],
            credentials["password"],
            f"{credentials['host']}:{credentials['port']}/{credentials['database']}",
            encoding="UTF-8",
        )

    elif credentials["db_type"] == "mssql":
        if credentials["user"] == "<WindowsUser>":
            # use local windows user via Windows Run As...
            # in a .bat / .cmd script use: runas /user:MyDomain\MyUser /savecred "cmd /K cd c:\myDir && python myScript.py"  # noqa: E501
            connection = pyodbc.connect(
                f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER=tcp:{credentials['host']},{credentials['port']};DATABASE={credentials['database']};Trusted_Connection=yes",  # noqa: E501
            )
        else:
            connection = pyodbc.connect(
                f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER=tcp:{credentials['host']},{credentials['port']};DATABASE={credentials['database']};UID={credentials['user']};PWD={credentials['password']}",  # noqa: E501
            )

    else:
        msg = f"Unsupported db_type: '{credentials['db_type']}'"
        raise ValueError(msg)

    cursor = connection.cursor()

    if verbose:
        print(
            f"connected to database {credentials['database']} on host {credentials['host']}",  # noqa: E501
        )

    return connection, cursor


def execute_sql(sql: str) -> list:
    """
    Execute SQL statement and return results as list.
    """
    sql_check_danger(sql=sql)
    results = []
    try:
        cursor.execute(sql)
        col_names = [desc[0] for desc in cursor.description]
        cnt_columns = len(col_names)
        results.append(col_names)
        cnt = 0
        for row in cursor:
            results.append(row)
            cnt += cnt_columns
            if cnt > cnt_max_cells:
                msg = (
                    f"WARNING: too many values, stopped after {cnt_max_cells} values",
                )
                raise Exception(msg)  # noqa: TRY301, TRY002
    except Exception as error:  # noqa: BLE001
        cursor.execute("ROLLBACK")
        print(error)
    return results


def sql_check_danger(sql: str) -> None:
    """
    Check SQl for bad code.

    raises Exception in case SQL contains dangerous commands
    Warning: incomplete, so not rely on this, use read only DB user too!
    """
    # bad = False
    bad_words = (
        "create",
        "alter",
        "drop",
        "grant",
        "revoke",
        "insert",
        "update",
        "delete",
        "truncate",
        "commit",
    )
    res = re.match(r"\b(" + "|".join(bad_words) + r")\b", sql.lower())
    if res is not None:
        # bad = True
        msg = f"ERROR: dangerous SQL: \n{sql}"
        raise Exception(msg)  # noqa: TRY002
        # sys.exit(1)
    res = re.match(r"[^\b](select)\b", sql.lower())
    if res is not None:
        # bad = True
        msg = f"ERROR: not valid SQL: \n{sql}"
        raise Exception(msg)  # noqa: TRY002
    # return bad


#
# 2. export functions
#


def sql2csv(results: list, filename: str) -> None:
    """
    Write results into csv file.
    """
    outfile = Path(filename + ".csv")
    with outfile.open(mode="w", encoding="utf-8", newline=csv_newline) as fh:
        csvwriter = csv.writer(
            fh,
            delimiter=csv_delimiter,
            quotechar=csv_quotechar,
        )
        col_names = results[0]  # header row
        csvwriter.writerow(col_names)
        for k in range(1, len(results)):
            row = results[k]
            row_str = []
            for value in row:
                value_str = _convert_value_to_string(
                    value,
                    remove_linebreaks=True,
                    remove_quotes=True,
                    trim=True,
                )
                row_str.append(value_str)
            csvwriter.writerow(row_str)


def sql2xlsx(results: list, filename: str) -> None:
    """
    Write results into Excel .xlsx file.
    """
    outfile = Path(filename + ".xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    col_names = results[0]  # header row
    i = 1
    j = 1
    for value in col_names:
        # note: excel index start here with 1
        cell = sheet.cell(row=i, column=j)  # type: ignore
        cell.value = value
        cell.font = openpyxl.styles.Font(bold=True)  # type: ignore

        j += 1

    # contents
    i = 2
    for k in range(1, len(results)):
        row = results[k]
        #    for row in results:
        j = 1
        for value in row:
            cell = sheet.cell(row=i, column=j)  # type: ignore
            cell.value = value
            j += 1
        i += 1

    # TODO: autosize columns width
    # V1 from https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
    # V2 # from https://izziswift.com/openpyxl-adjust-column-width-size/

    workbook.save(outfile)


def _convert_value_to_string(  # noqa: PLR0912, C901
    value,  # noqa: ANN001
    *,
    remove_linebreaks: bool = True,
    remove_quotes: bool = True,
    trim: bool = True,
) -> str:
    """
    Convert SQL field types to strings, used in sql2csv.
    """
    value_str = ""
    t = type(value)
    if value is None:
        value_str = ""
    elif t == str:
        value_str = value
        if remove_linebreaks:
            value_str = (
                value_str.replace("\r\n", " ")
                .replace(
                    "\n",
                    " ",
                )
                .replace("\r", " ")
            )
        if remove_quotes:
            value_str = value_str.replace("'", "").replace(
                '"',
                "",
            )  # remove " and '
        if trim:
            value_str = value_str.strip()
            value_str = value_str.replace("\t", " ")  # tabs
            value_str = value_str.replace("  ", " ").replace(
                "  ",
                " ",
            )  # multiple spaces

    elif t in (int, float):
        value_str = str(value)
    elif t == datetime.datetime:
        value_str = value.strftime(csv_format_datetime)
    elif t == datetime.date:
        value_str = value.strftime(csv_format_date)
    elif t == bool:
        if value is True:
            value_str = "1"
        if value is False:
            value_str = "0"
    elif t in (decimal.Decimal, datetime.timedelta):
        value_str = str(value)
    else:
        msg = f"unhandled column type: {t}"
        raise Exception(msg)  # noqa: TRY002
    return value_str


#
# 3. checksum functions
#


def gen_checksum(s: str, my_secret: str) -> str:
    """
    Calculate a sha256 checksum/hash of a string.

    add a "secret/salt" to the string to prevent others from being able to reproduce
    the checksum without knowing the secret
    """
    m = hashlib.sha256()
    m.update((s + my_secret).encode("utf-8"))
    return m.hexdigest()


def check_for_valid_hash_file(sql: str, filename: str) -> bool:
    """
    Check if file hash is valid.

    return False if .hash file is missing or contains a wrong hash
    """
    valid = True
    outfile = Path(filename + ".hash")
    if valid and not outfile.exists():
        valid = False
        print(f"ERROR: missing checksum file '{outfile}'")

    if valid:
        with outfile.open(
            encoding="utf-8",
            newline="\n",
        ) as fh:
            checksum_file = fh.read()
        checksum_calc = gen_checksum(s=sql, my_secret=hash_salt)
        if checksum_file != checksum_calc:
            valid = False
            print("ERROR: checksum mismatch")

    return valid


#
# 4. helper functions
#


def remove_old_output_files(filename: str) -> None:
    """
    Remove output files prior to re-creation.

    this is done prior to hash check to ensure that there is no output file in case
    the hash is bad
    """
    for ext in (".csv", ".xlsx"):
        outfile = Path(filename + ext)
        outfile.unlink(missing_ok=True)


#
# 5. main: loop over .sql files
#


if __name__ == "__main__":
    (connection, cursor) = connect()
    for filepath in Path().glob("*.sql"):
        print(f"File: {filepath}")
        (filename, file_ext) = (filepath.name, filepath.suffix)

        remove_old_output_files(filename)

        # not set newline type here, it might be \n or \r\n
        with filepath.open(encoding="utf-8") as fh:
            sql = fh.read()

        if hash_salt != "":  # perform hash check
            ret = check_for_valid_hash_file(
                sql=sql,
                filename=filename,
            )
            if ret is not True:
                continue

        # stop time to execute SQL
        start = datetime.datetime.now()  # noqa: DTZ005
        results = execute_sql(sql=sql)
        end = datetime.datetime.now()  # noqa: DTZ005
        duration = (end - start).total_seconds()
        print(f"Rows: {len(results)}")
        print(f"Duration: {duration:.3f}")
        if len(results) > 1:
            sql2csv(results=results, filename=filename)
            sql2xlsx(results=results, filename=filename)

    if connection:
        cursor.close()
        connection.close()
        print("Database connection closed")
