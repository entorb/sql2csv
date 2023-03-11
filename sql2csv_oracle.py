#!/usr/bin/python
"""
connects to a database
reads all .sql files of current directory
excecutes one after the other
results are written to .csv files

ONLY USE READ-ONLY DB-USER ACCOUNTS via:
GRANT SELECT ON ALL TABLES IN SCHEMA schema_name TO username

Requirements
Oracle Instant Client - Basic Light Package
from
https://www.oracle.com/database/technologies/instant-client/winx64-64-downloads.html
download and unzip and add dir to path

"""

# Convert to .exe via
# pyinstaller --onefile --console sql2csv_oracle.py

# FIXME: this security check is not working
# if (sql1.find('insert') or sql1.find('update') or sql1.find('delete') or sql1.find('drop') or sql1.find('grant')):
#     print ("invalid SQL")
#     sys.exit(1)


import os
import glob
import cx_Oracle
import datetime
import openpyxl

credentials_MyDB1 = {'host': 'myHost', 'port': 5432,
                     'database': 'myDB', 'user': 'myUser', 'password': 'myPwd'}

credentials = credentials_MyDB1


def connect():
    """ Connect to the PostgreSQL database server """
    connection = None
    cursor = None
    try:
        connection = cx_Oracle.connect(
            credentials['user'],
            credentials['password'],
            f"{credentials['host']}:{credentials['port']}/{credentials['database']}",
            encoding="UTF-8"
        )
        cursor = connection.cursor()
        print(
            f"connected to database {credentials['database']} on host {credentials['host']}")
    except (Exception) as error:
        print("Error while connecting to Oracle", error)

    return connection, cursor


def sql2csv(sql: str, outfile: str = 'out.csv'):
    """ Excecute SQL statement and write results into csv file """
    if os.path.isfile(outfile):
        os.remove(outfile)
    try:
        with open(outfile, mode='w', encoding='utf-8', newline='\n') as fh:
            cursor.execute(sql)
            colnames = [desc[0] for desc in cursor.description]
            fh.write("\t".join(colnames))
            fh.write("\n")
            for row in cursor:
                row_str = []
                for value in row:
                    value_str = convert_to_string(
                        value, remove_linebreaks=True, trim=True)
                    row_str.append(value_str)

                fh.write("\t".join(row_str))
                fh.write("\n")

    except (Exception) as error:
        cursor.execute('ROLLBACK')
        print(error)


def sql2xlsx(sql: str, outfile: str = 'out.xlsx'):
    """ Excecute SQL statement and write results into xlsx file """
    if os.path.isfile(outfile):
        os.remove(outfile)
    workbookOut = openpyxl.Workbook()
    sheetOut = workbookOut.active

    cursor.execute(sql)
    # header
    colnames = [desc[0] for desc in cursor.description]
    i = 1
    j = 1
    for value in colnames:
        # note: excel index start here with 1
        cellOut = sheetOut.cell(row=i, column=j)
        cellOut.value = value
        cellOut.font = openpyxl.styles.Font(bold=True)
        j += 1

    # contents
    i = 2
    for row in cursor:
        j = 1
        for value in row:
            cellOut = sheetOut.cell(row=i, column=j)
            cellOut.value = value
            j += 1
        i += 1

    # TODO: autosize columns width
    # V1 from https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
    # V2 # from https://izziswift.com/openpyxl-adjust-column-width-size/

    workbookOut.save(outfile)


def convert_to_string(value, remove_linebreaks: bool = True, trim: bool = True) -> str:
    value_str = ""
    t = type(value)
    if t == str:
        value_str = value
        if remove_linebreaks:
            value_str = value_str.replace("\r\n", " ").replace(
                "\n", " ").replace("\r", " ")
        if trim:
            value_str = value_str.strip()
            value_str = value_str.replace("\t", " ")  # tabs
            value_str = value_str.replace("  ", " ").replace(
                "  ", " ")  # multiple spaces

    elif t == int:
        value_str = str(value)
    elif t == datetime.datetime:
        value_str = value.strftime("%y%m%d-%H%M%S")
    else:
        print(f"unhandled column type: {t}")
    return value_str


if __name__ == '__main__':
    (connection, cursor) = connect()
    # sql = "SELECT SECUENCIA, TIPO, PARAM1, PARAM2, ESTADO, ERROR, ERROR_TXT, TMP_EMISION, TMP_TRATO, CREATION_DATE, LAST_UPDATE_DATE, LAST_UPDATE_PROCESS FROM COM_INBOX"
    for filename in glob.glob("*.sql"):
        print(f'File: {filename}')
        (fileBaseName, fileExtension) = os.path.splitext(filename)

        with open(filename, mode='r', encoding='utf-8') as fh:
            sql = fh.read()
        sql2csv(sql=sql, outfile=fileBaseName+".csv")
        sql2xlsx(sql=sql, outfile=fileBaseName+".xlsx")

    if (connection):
        cursor.close()
        connection.close()
        print("Oracle connection closed")
